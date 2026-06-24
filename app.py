# -*- coding: utf-8 -*-
"""
Painel web local (v2) para testar e operar o detector de placas.
Roda em 127.0.0.1 (so nesta maquina). Inicie com:  python app.py
(ou dois cliques em INICIAR_PAINEL.bat)

Arquitetura: a DETECCAO (lenta, CPU) roda uma vez e salva deteccoes.csv + recortes
locais. Tudo o mais (limiar, filtros, avaliacao, mosaico, sequencia, mapa, exportar)
recalcula na hora a partir desse cache -- por isso o slider responde ao vivo.
"""
import os, re, csv, json, time, glob, threading, subprocess, webbrowser, urllib.parse
import xml.etree.ElementTree as ET
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

RAIZ = os.path.dirname(os.path.abspath(__file__))
SAIDAS = os.path.join(RAIZ, "saidas")
DADOS = os.path.join(RAIZ, "dados")
OUTPUT = os.path.join(RAIZ, "output")   # fotos de referencia (quadro inteiro c/ a caixa)
ROOT_IMAGENS = r"\\192.168.0.210\Setores\Setor Dev\_TESTES\Placas"
PORTA = 8765

ESTADO = {
    "running": False, "done": False, "error": None, "step": "ocioso",
    "processed": 0, "total": 0, "found": 0, "log": [], "out": None,
    "t0": 0.0, "eta_s": 0, "elapsed_s": 0, "folders": [], "placas": [],
}
LOCK = threading.Lock()
PARAR = {"flag": False, "proc": None}   # controle do botao Parar
_folders_cache = None


def log(msg):
    with LOCK:
        ESTADO["log"].append(msg)
        ESTADO["log"] = ESTADO["log"][-80:]


# ------------------------------------------------------------------ pastas
def listar_pastas(root, max_dirs=800, max_result=120):
    global _folders_cache
    if _folders_cache is not None:
        return _folders_cache
    achadas, n = [], 0
    try:
        for dp, dn, fn in os.walk(root):
            n += 1
            if n > max_dirs:
                break
            if any(f.lower().endswith((".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp")) for f in fn):
                achadas.append(dp)
            if len(achadas) >= max_result:
                break
    except Exception as e:
        log(f"[aviso] listar pastas: {e}")
    _folders_cache = sorted(achadas)
    return _folders_cache


def listar_gabaritos():
    return sorted(glob.glob(os.path.join(DADOS, "*.csv")))


# ------------------------------------------------------------------ deteccao
def worker(folders, imgsz, limite):
    try:
        nome = os.path.basename(folders[0].rstrip("\\/")) or "pasta"
        out = os.path.join(SAIDAS, "PAINEL_" + re.sub(r"[^\w\-]", "_", nome))
        with LOCK:
            ESTADO.update({"out": out, "step": "detectando", "processed": 0,
                           "total": 0, "found": 0, "t0": time.time(),
                           "folders": folders, "placas": []})
        args = ["python", os.path.join(RAIZ, "src", "detectar.py"),
                "--pastas", ";".join(folders), "--stride", "1",
                "--conf", "0.04", "--imgsz", str(imgsz), "--out", out]
        if limite > 0:
            args += ["--max", str(limite)]
        env = dict(os.environ, PYTHONUNBUFFERED="1", YOLO_VERBOSE="False",
                   PYTHONIOENCODING="utf-8")
        p = subprocess.Popen(args, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                             text=True, encoding="utf-8", bufsize=1, env=env, cwd=RAIZ)
        PARAR["proc"] = p
        ultima = ""
        for linha in p.stdout:
            linha = linha.strip()
            if not linha:
                continue
            ultima = linha
            m = re.search(r"(\d+) imagens a processar", linha)
            if m:
                with LOCK:
                    ESTADO["total"] = int(m.group(1))
            m = re.search(r"\[prog\] (\d+)/(\d+)", linha)
            if m:
                proc, tot = int(m.group(1)), int(m.group(2))
                el = time.time() - ESTADO["t0"]
                eta = (el / proc) * (tot - proc) if proc else 0
                with LOCK:
                    ESTADO.update({"processed": proc, "total": tot,
                                   "elapsed_s": int(el), "eta_s": int(eta)})
            m = re.search(r"deteccoes:\s*(\d+)", linha)
            if m:
                with LOCK:
                    ESTADO["found"] = int(m.group(1))
            if linha[:6] in ("[prog]", "[info]") or linha.startswith("[ok]"):
                log(linha)
        p.wait()
        if PARAR["flag"]:
            with LOCK:
                ESTADO.update({"step": "interrompido", "running": False, "done": False,
                               "eta_s": 0})
            log("[parado] Processo interrompido pelo usuario.")
            return
        if p.returncode != 0:
            msg = ultima.replace("[erro]", "").strip() or "deteccao terminou com erro"
            if "nenhuma imagem" in msg.lower():
                msg = "Nenhuma imagem encontrada nessa pasta (formatos aceitos: jpg, jpeg, png, bmp, tif, webp)."
            raise RuntimeError(msg)
        with LOCK:
            ESTADO.update({"step": "deteccao pronta", "done": True, "running": False,
                           "processed": ESTADO["total"] or ESTADO["processed"],
                           "elapsed_s": int(time.time() - ESTADO["t0"]), "eta_s": 0})
        log("[ok] Deteccao concluida. Ajuste os filtros e veja o resultado ao vivo.")
    except Exception as e:
        with LOCK:
            ESTADO.update({"error": str(e), "running": False, "step": "erro"})
        log(f"[ERRO] {e}")


# ------------------------------------------------------------------ dedup (em memoria, instantaneo)
def ler_deteccoes(out):
    fp = os.path.join(out, "deteccoes.csv")
    if not os.path.exists(fp):
        return []
    with open(fp, encoding="utf-8-sig") as f:
        dets = list(csv.DictReader(f))
    for d in dets:
        d["km"] = float(d["km"]); d["conf"] = float(d["conf"])
        d["area_frac"] = float(d["area_frac"]); d["completa"] = int(d["completa"])
        h = max(1.0, float(d["y2"]) - float(d["y1"]))
        d["aspecto"] = (float(d["x2"]) - float(d["x1"])) / h
    return dets


def deduplicar(dets, conf, min_quadros, min_asp, max_asp, min_area, janela_km, lado_filtro):
    # filtros basicos
    sel = [d for d in dets if d["conf"] >= conf and min_asp <= d["aspecto"] <= max_asp
           and d["area_frac"] >= min_area]
    for d in sel:
        d["cx"] = (float(d["x1"]) + float(d["x2"])) / 2.0   # centro horizontal (px)

    # RASTREAMENTO: cada placa fisica segue uma trajetoria propria (a caixa anda na
    # horizontal e cresce conforme o veiculo se aproxima). Agrupar por proximidade de km
    # juntava placas diferentes; aqui agrupamos por CONTINUIDADE de posicao entre quadros.
    MAXGAP = 0.03    # km (~30 m): tolera ate ~3 quadros sem deteccao no mesmo objeto
    XTOL = 360.0     # px: deslocamento horizontal maximo entre quadros p/ ser a MESMA placa
    grupos = []
    for lado in ("E", "D"):
        if lado_filtro in ("E", "D") and lado != lado_filtro:
            continue
        seq = sorted([d for d in sel if d["lado_img"] == lado], key=lambda x: x["km"])
        tracks = []   # {"dets", "last_km", "last_cx", "closed"}
        for d in seq:
            melhor_t, melhor_dx = None, XTOL + 1
            for t in tracks:
                if t["closed"]:
                    continue
                if d["km"] - t["last_km"] > MAXGAP:   # objeto ficou para tras -> fecha
                    t["closed"] = True
                    continue
                # a MESMA placa cresce ao se aproximar; se encolheu muito, e' outra placa
                if d["area_frac"] < t["last_area"] * 0.55:
                    continue
                dx = abs(d["cx"] - t["last_cx"])
                if dx < melhor_dx:
                    melhor_dx, melhor_t = dx, t
            if melhor_t is not None and melhor_dx <= XTOL:
                melhor_t["dets"].append(d)
                melhor_t["last_km"] = d["km"]
                melhor_t["last_cx"] = d["cx"]
                melhor_t["last_area"] = d["area_frac"]
            else:
                tracks.append({"dets": [d], "last_km": d["km"], "last_cx": d["cx"],
                               "last_area": d["area_frac"], "closed": False})
        grupos.extend(t["dets"] for t in tracks)
    grupos = [g for g in grupos if len(g) >= min_quadros]
    placas = []
    for g in grupos:
        melhor = sorted(g, key=lambda x: (x["completa"], x["area_frac"]))[-1]
        # membros = os quadros EXATOS deste agrupamento (a mesma placa, sem misturar vizinhas)
        membros = [{"km": round(d["km"], 3), "conf": round(d["conf"], 3),
                    "area": round(d["area_frac"], 5), "completa": d["completa"],
                    "crop": d.get("crop", "")} for d in sorted(g, key=lambda x: x["km"])]
        placas.append({
            "km": round(melhor["km"], 3), "lado": melhor["lado_img"],
            "conf": round(melhor["conf"], 4), "area_frac": round(melhor["area_frac"], 6),
            "completa": melhor["completa"], "n_quadros": len(g),
            "x1": melhor["x1"], "y1": melhor["y1"], "x2": melhor["x2"], "y2": melhor["y2"],
            "imagem": melhor["imagem"], "crop": melhor.get("crop", ""),
            "membros": membros,
        })
    placas.sort(key=lambda x: x["km"])
    return placas


def salvar_placas(out, placas):
    fp = os.path.join(out, "placas_unicas.csv")
    with open(fp, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["km", "lado", "conf", "area_frac", "completa",
                                          "n_quadros", "x1", "y1", "x2", "y2", "imagem", "crop"],
                           extrasaction="ignore")
        w.writeheader(); w.writerows(placas)
    return fp


def _cv2_ler(caminho):
    """cv2.imread falha silenciosamente com caminhos Unicode no Windows; usa fromfile."""
    import cv2, numpy as np
    try:
        arr = np.fromfile(caminho, dtype=np.uint8)
        return cv2.imdecode(arr, cv2.IMREAD_COLOR)
    except Exception:
        return None


def _cv2_salvar(caminho, img):
    """cv2.imwrite também falha com caminhos Unicode; usa imencode + tofile."""
    import cv2, numpy as np
    try:
        ok, buf = cv2.imencode(".jpg", img)
        if ok:
            np.array(buf).tofile(caminho)
            return True
    except Exception:
        pass
    return False


def gerar_referencia(out, placa, i):
    caminho_orig = placa["imagem"]
    img = _cv2_ler(caminho_orig)
    if img is None:
        nome = os.path.basename(caminho_orig)
        for pasta in (ESTADO.get("folders") or []):
            img = _cv2_ler(os.path.join(pasta, nome))
            if img is not None:
                break
    if img is None:
        return None
    import cv2
    x1, y1 = int(float(placa["x1"])), int(float(placa["y1"]))
    x2, y2 = int(float(placa["x2"])), int(float(placa["y2"]))
    cv2.rectangle(img, (x1, y1), (x2, y2), (0, 220, 0), 5)
    rot = f"#{i+1}  km {float(placa['km']):.3f}  lado {placa['lado']}  conf {float(placa['conf']):.2f}"
    cv2.rectangle(img, (0, 0), (min(img.shape[1], 980), 48), (0, 0, 0), -1)
    cv2.putText(img, rot, (12, 33), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2)
    run = os.path.basename(out.rstrip("\\/"))
    odir = os.path.join(OUTPUT, run)
    os.makedirs(odir, exist_ok=True)
    nome = f"placa_{i+1:03d}_km{float(placa['km']):.3f}_{placa['lado']}.jpg".replace(" ", "_")
    cam = os.path.join(odir, nome)
    _cv2_salvar(cam, img)
    return cam


# ------------------------------------------------------------------ avaliacao vs gabarito
def avaliar(placas, gabarito_path, tol_m, usar_lado):
    if not gabarito_path or not os.path.exists(gabarito_path):
        return None
    with open(gabarito_path, encoding="utf-8-sig") as f:
        gt = list(csv.DictReader(f))
    for g in gt:
        try:
            g["_km"] = float(g.get("km_abs") or g.get("km"))
        except Exception:
            g["_km"] = None
    gt = [g for g in gt if g["_km"] is not None]
    if not placas or not gt:
        return {"erro": "sem dados"}
    tol = tol_m / 1000.0
    dlo, dhi = min(p["km"] for p in placas), max(p["km"] for p in placas)
    glo, ghi = min(g["_km"] for g in gt), max(g["_km"] for g in gt)
    ov_lo, ov_hi = max(dlo, glo), min(dhi, ghi)
    if ov_hi <= ov_lo:
        return {"sem_overlap": True, "img_km": [round(dlo, 1), round(dhi, 1)],
                "gab_km": [round(glo, 1), round(ghi, 1)]}
    gtf = [g for g in gt if ov_lo - tol <= g["_km"] <= ov_hi + tol]
    usados = set(); achadas = 0; faltaram = []
    for g in gtf:
        bi, bd = None, 1e9
        for i, p in enumerate(placas):
            if i in usados:
                continue
            if usar_lado and p.get("lado") and g.get("lado") and p["lado"] != g["lado"]:
                continue
            dd = abs(p["km"] - g["_km"])
            if dd < bd:
                bd, bi = dd, i
        if bi is not None and bd <= tol:
            usados.add(bi); achadas += 1
        else:
            faltaram.append({"km": round(g["_km"], 3), "lado": g.get("lado", ""),
                             "codigo": g.get("codigo", "")})
    falsos = [i for i in range(len(placas)) if i not in usados]
    n_gt, n_det = len(gtf), len(placas)
    return {
        "n_gt": n_gt, "n_det": n_det, "achadas": achadas, "fp": len(falsos),
        "recall": round(100 * achadas / n_gt, 1) if n_gt else 0,
        "precisao": round(100 * len(usados) / n_det, 1) if n_det else 0,
        "fp_rel": round(100 * len(falsos) / n_gt, 1) if n_gt else 0,
        "faltaram": faltaram[:60], "falsos_idx": falsos,
        "overlap": [round(ov_lo, 1), round(ov_hi, 1)],
    }


# ------------------------------------------------------------------ GPX -> mapa
def parse_gpx_pasta(folders):
    pts = []
    for folder in folders:
        pai = os.path.dirname(folder)
        base = os.path.basename(folder.rstrip("\\/"))
        cands = glob.glob(os.path.join(pai, "*.GPX")) + glob.glob(os.path.join(pai, "*.gpx"))
        alvo = None
        for c in cands:
            if os.path.splitext(os.path.basename(c))[0].lower() == base.lower():
                alvo = c; break
        if not alvo and cands:
            alvo = cands[0]
        if not alvo:
            continue
        try:
            tree = ET.parse(alvo)
            for el in tree.iter():
                if el.tag.endswith("trkpt") or el.tag.endswith("rtept") or el.tag.endswith("wpt"):
                    la, lo = el.get("lat"), el.get("lon")
                    if la and lo:
                        pts.append([float(la), float(lo)])
        except Exception as e:
            log(f"[aviso] gpx {alvo}: {e}")
    return pts


# ------------------------------------------------------------------ servidor
def dentro(base, caminho):
    return os.path.realpath(caminho).startswith(os.path.realpath(base))


class H(BaseHTTPRequestHandler):
    def log_message(self, *a):
        pass

    def _send(self, code, ctype, body):
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def _json(self, obj, code=200):
        self._send(code, "application/json; charset=utf-8",
                   json.dumps(obj, ensure_ascii=False).encode("utf-8"))

    def _file_img(self, path, base=None):
        base = base or SAIDAS
        if not path or not os.path.exists(path) or not dentro(base, path):
            self._send(404, "text/plain", b"sem imagem"); return
        with open(path, "rb") as f:
            self._send(200, "image/jpeg", f.read())

    def _static(self, path):
        base = os.path.join(RAIZ, "web")
        if not os.path.isfile(path) or not dentro(base, path):
            self._send(404, "text/plain", b"nao encontrado"); return
        ext = os.path.splitext(path)[1].lower()
        ct = {".css": "text/css", ".js": "application/javascript", ".png": "image/png",
              ".html": "text/html; charset=utf-8"}.get(ext, "application/octet-stream")
        with open(path, "rb") as f:
            self._send(200, ct, f.read())

    # -------------------- GET
    def do_GET(self):
        u = urllib.parse.urlparse(self.path)
        q = urllib.parse.parse_qs(u.query)
        out = ESTADO.get("out")
        if u.path == "/favicon.ico":
            self._send(200, "image/x-icon", b"")
        elif u.path.startswith("/vendor/"):
            self._static(os.path.join(RAIZ, "web", "vendor", os.path.basename(u.path)))
        elif u.path == "/":
            with open(os.path.join(RAIZ, "web", "index.html"), "rb") as f:
                self._send(200, "text/html; charset=utf-8", f.read())
        elif u.path == "/api/folders":
            self._json({"root": ROOT_IMAGENS, "pastas": listar_pastas(ROOT_IMAGENS),
                        "gabaritos": listar_gabaritos()})
        elif u.path == "/api/status":
            with LOCK:
                s = {k: ESTADO[k] for k in ("running", "done", "error", "step",
                     "processed", "total", "found", "eta_s", "elapsed_s")}
                s["log"] = ESTADO["log"][-20:]
            self._json(s)
        elif u.path == "/api/latest":
            d = os.path.join(out, "anotadas") if out else None
            arqs = ([os.path.join(d, x) for x in os.listdir(d) if x.lower().endswith(".jpg")]
                    if d and os.path.isdir(d) else [])
            self._file_img(max(arqs, key=os.path.getmtime) if arqs else None)
        elif u.path == "/api/mosaico":
            if out:
                uni = os.path.join(out, "placas_unicas.csv")
                if os.path.exists(uni):
                    subprocess.run(["python", os.path.join(RAIZ, "src", "montagem.py"),
                                    uni, os.path.join(out, "mosaico"), "--cols", "7",
                                    "--tile", "200"], capture_output=True, cwd=RAIZ)
                self._file_img(os.path.join(out, "mosaico", "mosaico_placas.jpg"))
            else:
                self._send(404, "text/plain", b"sem run")
        elif u.path == "/api/file":
            self._file_img((q.get("path", [""])[0]))
        elif u.path == "/api/referencia":
            i = int(q.get("i", [-1])[0])
            placas = ESTADO.get("placas") or []
            if out and 0 <= i < len(placas):
                self._file_img(gerar_referencia(out, placas[i], i), base=OUTPUT)
            else:
                self._send(404, "text/plain", b"sem placa")
        elif u.path == "/api/sequence":
            # retorna EXATAMENTE os quadros do agrupamento da placa de indice i
            i = int(q.get("i", [-1])[0])
            placas = ESTADO.get("placas") or []
            seq = placas[i].get("membros", []) if 0 <= i < len(placas) else []
            self._json({"seq": seq})
        elif u.path == "/api/mapa":
            pts = parse_gpx_pasta(ESTADO.get("folders") or [])
            placas = ESTADO.get("placas") or []
            marc = []
            if pts and placas:
                kms = [p["km"] for p in placas]
                kmin, kmax = min(kms), max(kms)
                rng = (kmax - kmin) or 1
                for i, p in enumerate(placas):
                    frac = (p["km"] - kmin) / rng
                    idx = min(len(pts) - 1, max(0, int(frac * (len(pts) - 1))))
                    marc.append({"i": i, "km": p["km"], "lado": p["lado"],
                                 "lat": pts[idx][0], "lon": pts[idx][1]})
            self._json({"track": pts[::5], "placas": marc})
        elif u.path == "/api/picker":
            try:
                import tkinter as tk
                from tkinter import filedialog
                root = tk.Tk()
                root.withdraw()
                root.wm_attributes("-topmost", 1)
                pasta = filedialog.askdirectory(title="Selecionar pasta com imagens")
                root.destroy()
                self._json({"path": pasta or ""})
            except Exception as e:
                self._json({"path": "", "erro": str(e)})
        else:
            self._send(404, "text/plain", b"nao encontrado")

    # -------------------- POST
    def do_POST(self):
        u = urllib.parse.urlparse(self.path)
        n = int(self.headers.get("Content-Length", 0))
        body = json.loads(self.rfile.read(n) or b"{}") if n else {}
        if u.path == "/api/run":
            brutas = [(f or "").strip().strip('"').strip("'") for f in (body.get("folders") or [])]
            brutas = [f for f in brutas if f]
            folders = [f for f in brutas if os.path.isdir(f)]
            if not folders:
                alvo = " | ".join(brutas) if brutas else "(nenhuma)"
                self._json({"ok": False, "msg": "Pasta nao encontrada: " + alvo}, 400); return
            if ESTADO["running"]:
                self._json({"ok": False, "msg": "ja esta rodando"}, 409); return
            imgsz = int(body.get("imgsz") or 1280); limite = int(body.get("limite") or 0)
            PARAR["flag"] = False; PARAR["proc"] = None
            with LOCK:
                ESTADO.update({"running": True, "done": False, "error": None,
                               "step": "iniciando", "processed": 0, "total": 0,
                               "found": 0, "log": [], "placas": []})
            threading.Thread(target=worker, args=(folders, imgsz, limite), daemon=True).start()
            self._json({"ok": True})
        elif u.path == "/api/stop":
            PARAR["flag"] = True
            pr = PARAR.get("proc")
            if pr and pr.poll() is None:
                try:
                    pr.terminate()
                except Exception:
                    pass
            with LOCK:
                ESTADO.update({"running": False, "step": "interrompido", "eta_s": 0})
            self._json({"ok": True})
        elif u.path == "/api/refine":
            out = ESTADO.get("out")
            if not out:
                self._json({"ok": False, "msg": "rode a deteccao primeiro"}, 400); return
            dets = ler_deteccoes(out)
            placas = deduplicar(
                dets, float(body.get("conf", 0.10)), int(body.get("min_quadros", 1)),
                float(body.get("min_aspecto", 0.45)), float(body.get("max_aspecto", 4.0)),
                float(body.get("min_area", 0.0002)), float(body.get("janela_km", 0.06)),
                body.get("lado", "ambos"))
            salvar_placas(out, placas)
            with LOCK:
                ESTADO["placas"] = placas
            ev = avaliar(placas, body.get("gabarito", ""), float(body.get("tol_m", 60)),
                         bool(body.get("usar_lado", False)))
            slim = [{k: v for k, v in p.items() if k != "membros"} for p in placas]
            self._json({"ok": True, "n": len(placas), "placas": slim, "eval": ev,
                        "n_det": len(dets)})
        elif u.path == "/api/export":
            out = ESTADO.get("out")
            placas = ESTADO.get("placas") or []
            rej = set(body.get("rejeitados") or [])
            sel = [p for i, p in enumerate(placas) if i not in rej]
            fmt = body.get("formato", "csv")
            try:
                if fmt == "xlsx":
                    cam = exportar_xlsx(out, sel)
                else:
                    cam = exportar_csv(out, sel)
                self._json({"ok": True, "path": cam, "n": len(sel)})
            except Exception as e:
                self._json({"ok": False, "msg": str(e)}, 500)
        elif u.path == "/api/salvar_referencias":
            out = ESTADO.get("out"); placas = ESTADO.get("placas") or []
            if not out or not placas:
                self._json({"ok": False, "msg": "rode a deteccao primeiro"}, 400); return
            n, pasta = 0, ""
            for i, p in enumerate(placas):
                c = gerar_referencia(out, p, i)
                if c:
                    n += 1; pasta = os.path.dirname(c)
            self._json({"ok": True, "n": n, "path": pasta})
        else:
            self._send(404, "text/plain", b"nao encontrado")


def exportar_csv(out, placas):
    cam = os.path.join(out, "INVENTARIO_detectado.csv")
    with open(cam, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["N", "km", "+m", "Lado", "Confianca", "Completa", "Foto"])
        for i, p in enumerate(placas, 1):
            kmint = int(p["km"]); m = int(round((p["km"] - kmint) * 1000))
            w.writerow([i, kmint, m, p["lado"], f"{p['conf']:.2f}",
                        "sim" if p["completa"] else "nao", os.path.basename(p.get("crop", ""))])
    return cam


def exportar_xlsx(out, placas):
    import openpyxl
    from openpyxl.drawing.image import Image as XLImg
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Sinalizacao detectada"
    ws.append(["N", "Foto", "km", "+m", "Lado", "Confianca", "Completa"])
    ws.column_dimensions["B"].width = 16
    for i, p in enumerate(placas, 1):
        r = i + 1
        kmint = int(p["km"]); m = int(round((p["km"] - kmint) * 1000))
        ws.cell(r, 1, i); ws.cell(r, 3, kmint); ws.cell(r, 4, m)
        ws.cell(r, 5, p["lado"]); ws.cell(r, 6, round(p["conf"], 2))
        ws.cell(r, 7, "sim" if p["completa"] else "nao")
        cp = p.get("crop", "")
        if cp and os.path.exists(cp):
            try:
                im = XLImg(cp); im.width = 90; im.height = 90
                ws.row_dimensions[r].height = 70
                ws.add_image(im, f"B{r}")
            except Exception:
                pass
    cam = os.path.join(out, "INVENTARIO_detectado.xlsx")
    wb.save(cam)
    return cam


def main():
    os.makedirs(SAIDAS, exist_ok=True)
    srv = ThreadingHTTPServer(("127.0.0.1", PORTA), H)
    url = f"http://127.0.0.1:{PORTA}/"
    print(f"\n  Painel de placas em: {url}")
    print("  (deixe esta janela aberta; feche para parar)\n")
    if not os.environ.get("PLACAS_NO_BROWSER"):
        try:
            webbrowser.open(url)
        except Exception:
            pass
    srv.serve_forever()


if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
Gera um xlsx de avaliacao de uma RODADA de deteccao (saidas/PAINEL_*).

Cruza as PLACAS detectadas (placas_unicas.csv) com:
  - o GABARITO -> recall, faltaram, falsos. Aceita dois formatos:
      .xlsx (Comparacao.xlsx do outro projeto, PADRAO): aba 'listagem_arquivos',
             1 placa real por imagem com Manual=1; o km vem do nome da imagem.
      .csv  (inventario): 1 linha por placa (km_abs/km, lado, codigo).
  - a CONFERENCIA manual (conferencia.json: cada placa = 'real' (OK) / 'falsa') -> precisao

Saida com 3 abas:
  comparacao : 1 linha por placa do gabarito -> codigo/imagem, km, casou? km_ia, dist, ✓/✗
               (placas detectadas FORA do gabarito entram no fim como 'falso positivo')
  placas     : todas as placas detectadas -> km, lado, conf, n_quadros, conferencia, casou
  resumo     : totais + recall, precisao (gabarito), precisao (conferencia), F1

Uso:
  python gerar_relatorio_xlsx.py --rodada "saidas\\PAINEL_minha_pasta"
  python gerar_relatorio_xlsx.py --rodada "saidas\\PAINEL_x" \
         --gabarito "dados\\Comparacao.xlsx" --tol-m 30 --output "Relatorio.xlsx"

Obs.: o km da deteccao vem do nome do arquivo (km absoluto). Se o resumo mostrar
0 placas na faixa do gabarito, o gabarito nao cobre o trecho/estrada das fotos
(ex.: gabarito do MT-361 com fotos da BR-010) -- use o gabarito certo.
"""
import os
import re
import csv
import json
import argparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


RAIZ = os.path.dirname(os.path.abspath(__file__))
# gabarito padrao = marcacao manual do outro projeto (Comparacao.xlsx, aba
# 'listagem_arquivos', col Name + Manual). Cada imagem com Manual=1 e' uma placa real.
GABARITO_PADRAO = os.path.join(RAIZ, "dados", "Comparacao.xlsx")


def f(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def crop_key(caminho):
    """Mesma chave usada pelo painel: so o nome do arquivo do recorte."""
    return re.split(r"[\\/]", str(caminho or ""))[-1]


def km_do_nome(nome):
    """km absoluto = ultimo numero decimal do nome do arquivo (igual ao detector)."""
    ms = re.findall(r"\d+\.\d+", str(nome or ""))
    return float(ms[-1]) if ms else None


def carregar_placas(rodada):
    fp = os.path.join(rodada, "placas_unicas.csv")
    if not os.path.isfile(fp):
        raise SystemExit(f"[erro] nao achei placas_unicas.csv em: {rodada}")
    placas = []
    with open(fp, encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            placas.append({
                "km": f(r.get("km")),
                "lado": (r.get("lado") or "").strip(),
                "conf": f(r.get("conf")),
                "n_quadros": int(f(r.get("n_quadros"))),
                "completa": int(f(r.get("completa"))),
                "imagem": r.get("imagem") or "",
                "crop": r.get("crop") or "",
            })
    placas.sort(key=lambda p: (p["km"], 0 if p["lado"] == "E" else 1))
    return placas


def carregar_conferencia(rodada):
    """Retorna (nome, {crop_key: 'real'|'falsa'})."""
    fp = os.path.join(rodada, "conferencia.json")
    if not os.path.isfile(fp):
        return "", {}
    try:
        with open(fp, encoding="utf-8") as fh:
            d = json.load(fh)
        return d.get("nome") or "", d.get("marks") or {}
    except Exception:
        return "", {}


def carregar_gabarito(caminho, km_col="km_abs", sheet="listagem_arquivos",
                      header_row=1, col_name="Name", col_manual="Manual"):
    """Le o gabarito. Suporta dois formatos:
      .xlsx (Comparacao.xlsx) -> 1 placa real por imagem com Manual=1 (km vem do nome)
      .csv  (inventario)      -> 1 linha por placa (km_abs/km, lado, codigo)"""
    if not os.path.isfile(caminho):
        raise SystemExit(f"[erro] gabarito nao encontrado: {caminho}")
    if caminho.lower().endswith((".xlsx", ".xlsm")):
        return _gabarito_xlsx(caminho, sheet, header_row, col_name, col_manual)
    gt = []
    with open(caminho, encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            km = r.get(km_col)
            if km in (None, ""):
                km = r.get("km_abs") or r.get("km")
            kmv = f(km, None) if km not in (None, "") else None
            if kmv is None:
                continue
            gt.append({"km": kmv, "lado": (r.get("lado") or "").strip(),
                       "codigo": (r.get("codigo") or "").strip()})
    return gt


def _gabarito_xlsx(caminho, sheet, header_row, col_name, col_manual):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if header_row >= len(rows):
        raise SystemExit(f"[erro] gabarito sem cabecalho na linha {header_row}: {caminho}")
    hdr = [str(h).strip() if h is not None else "" for h in rows[header_row]]
    idx = {h: i for i, h in enumerate(hdr)}
    iN, iM = idx.get(col_name), idx.get(col_manual)
    if iN is None or iM is None:
        raise SystemExit(f"[erro] gabarito sem colunas '{col_name}'/'{col_manual}' "
                         f"na aba '{ws.title}' (cabecalho: {hdr})")
    gt = []
    for r in rows[header_row + 1:]:
        if iN >= len(r) or r[iN] in (None, ""):
            continue
        manual = r[iM] if iM < len(r) else None
        try:
            if int(f(manual)) != 1:
                continue
        except (TypeError, ValueError):
            continue
        km = km_do_nome(r[iN])
        if km is None:
            continue
        gt.append({"km": km, "lado": "", "codigo": crop_key(r[iN])})
    return gt


def comparar(placas, gt, tol_m, usar_lado):
    """Casa cada placa do gabarito com a placa detectada mais proxima (guloso, 1-p/-1).
    Igual a logica do painel (app.comparar): restringe ao trecho com sobreposicao."""
    tol = tol_m / 1000.0
    if not placas or not gt:
        return {"erro": "sem dados", "tol": tol}
    dlo, dhi = min(p["km"] for p in placas), max(p["km"] for p in placas)
    glo, ghi = min(g["km"] for g in gt), max(g["km"] for g in gt)
    ov_lo, ov_hi = max(dlo, glo), min(dhi, ghi)
    gtf = [g for g in gt if ov_lo - tol <= g["km"] <= ov_hi + tol] if ov_hi > ov_lo else []

    usados = {}                # idx placa -> codigo do gabarito que casou
    casados_gt = []            # (g, idx_placa|None, dist)
    for g in sorted(gtf, key=lambda x: x["km"]):
        bi, bd = None, 1e9
        for i, p in enumerate(placas):
            if i in usados:
                continue
            if usar_lado and p["lado"] and g["lado"] and p["lado"] != g["lado"]:
                continue
            dd = abs(p["km"] - g["km"])
            if dd < bd:
                bd, bi = dd, i
        if bi is not None and bd <= tol:
            usados[bi] = g["codigo"]
            casados_gt.append((g, bi, bd))
        else:
            casados_gt.append((g, None, None))
    return {
        "tol": tol, "n_gt": len(gtf), "n_det": len(placas),
        "achadas": sum(1 for _, i, _ in casados_gt if i is not None),
        "casados_gt": casados_gt, "usados": usados,
        "ov": (round(dlo, 3), round(dhi, 3), round(glo, 3), round(ghi, 3)),
    }


# ----------------------------------------------------------------- estilo xlsx
CAB_FONT = Font(bold=True, color="FFFFFF")
CAB_FILL = PatternFill("solid", fgColor="305496")
CORES = {"detectou": "C6EFCE", "nao detectou": "FFC7CE",
         "falso positivo": "FFEB9C", "real": "C6EFCE", "falsa": "FFC7CE"}


def cabecalho(ws, headers):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = CAB_FONT
        cell.fill = CAB_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def marca_humana(marks, placa):
    m = marks.get(crop_key(placa["crop"]), "")
    return {"real": "✓ real", "falsa": "✗ falsa"}.get(m, "")


def gerar_relatorio(rodada, gabarito=GABARITO_PADRAO, km_col="km_abs",
                    tol_m=30.0, usar_lado=True, output=""):
    """Gera o xlsx da rodada e devolve um dict com o caminho de saida + as metricas.
    Reutilizavel pelo CLI (main) e pelo painel web (app.py /api/relatorio)."""
    rodada = os.path.abspath(rodada)
    placas = carregar_placas(rodada)
    nome, marks = carregar_conferencia(rodada)
    gt = carregar_gabarito(gabarito, km_col)
    cmp = comparar(placas, gt, tol_m, usar_lado=usar_lado)

    nome_run = nome or os.path.basename(rodada)
    output = output or os.path.join(
        rodada, "Relatorio_" + re.sub(r"[^\w\-]", "_", nome_run) + ".xlsx")

    casados_gt = cmp.get("casados_gt", [])
    usados = cmp.get("usados", {})
    achadas = cmp.get("achadas", 0)
    n_gt = cmp.get("n_gt", 0)
    n_det = len(placas)
    faltaram = sum(1 for _, i, _ in casados_gt if i is None)
    falsos_idx = [i for i in range(n_det) if i not in usados]
    n_falso = len(falsos_idx)

    # conferencia manual
    n_real = sum(1 for p in placas if marks.get(crop_key(p["crop"])) == "real")
    n_falsa = sum(1 for p in placas if marks.get(crop_key(p["crop"])) == "falsa")
    n_conf = n_real + n_falsa

    # metricas
    recall = achadas / n_gt if n_gt else 0
    prec_gab = achadas / n_det if n_det else 0
    f1 = 2 * prec_gab * recall / (prec_gab + recall) if (prec_gab + recall) else 0
    prec_conf = n_real / n_conf if n_conf else 0

    # ---------------------------------------------------------------- workbook
    wb = openpyxl.Workbook()

    # aba 1: comparacao por placa do gabarito
    ws1 = wb.active
    ws1.title = "comparacao"
    cabecalho(ws1, ["codigo_gab", "km_gab", "lado_gab", "km_ia", "lado_ia",
                    "conf_ia", "dist_m", "conferencia", "resultado"])
    for g, i, dist in casados_gt:
        if i is None:
            ws1.append([g["codigo"], round(g["km"], 3), g["lado"], "", "", "", "",
                        "", "nao detectou"])
            res_cell = "nao detectou"
        else:
            p = placas[i]
            ws1.append([g["codigo"], round(g["km"], 3), g["lado"],
                        round(p["km"], 3), p["lado"], round(p["conf"], 3),
                        round(dist * 1000, 1), marca_humana(marks, p), "detectou"])
            res_cell = "detectou"
        cor = CORES.get(res_cell)
        if cor:
            ws1.cell(ws1.max_row, 9).fill = PatternFill("solid", fgColor=cor)
    # placas detectadas fora do gabarito (falsos positivos) no fim
    for i in falsos_idx:
        p = placas[i]
        ws1.append(["", "", "", round(p["km"], 3), p["lado"], round(p["conf"], 3),
                    "", marca_humana(marks, p), "falso positivo"])
        ws1.cell(ws1.max_row, 9).fill = PatternFill("solid", fgColor=CORES["falso positivo"])
    for col, w in zip("ABCDEFGHI", (40, 11, 9, 11, 8, 9, 9, 12, 16)):
        ws1.column_dimensions[col].width = w

    # aba 2: todas as placas detectadas
    ws2 = wb.create_sheet("placas")
    cabecalho(ws2, ["#", "km", "lado", "conf", "n_quadros", "completa",
                    "conferencia", "casou_gabarito", "crop"])
    for i, p in enumerate(placas):
        ws2.append([i + 1, round(p["km"], 3), p["lado"], round(p["conf"], 3),
                    p["n_quadros"], p["completa"], marca_humana(marks, p),
                    usados.get(i, ""), crop_key(p["crop"])])
        mk = marks.get(crop_key(p["crop"]))
        if mk in CORES:
            ws2.cell(ws2.max_row, 7).fill = PatternFill("solid", fgColor=CORES[mk])
    for col, w in zip("ABCDEFGHI", (5, 11, 6, 8, 11, 9, 12, 40, 16)):
        ws2.column_dimensions[col].width = w

    # aba 3: resumo
    ws3 = wb.create_sheet("resumo")
    cabecalho(ws3, ["metrica", "valor"])
    ov = cmp.get("ov")
    resumo = [
        ("Rodada", nome_run),
        ("Gabarito", os.path.basename(gabarito)),
        ("Tolerancia de casamento (m)", tol_m),
        ("Placas detectadas (IA)", n_det),
        ("Placas do gabarito (na faixa)", n_gt),
        ("Detectou (casou gabarito)", achadas),
        ("Nao detectou (faltaram)", faltaram),
        ("Falso positivo (fora do gabarito)", n_falso),
        ("Conferidas manualmente", n_conf),
        ("  ✓ real", n_real),
        ("  ✗ falsa", n_falsa),
        ("Recall (gabarito detectado)", f"{recall:.1%}"),
        ("Precisao vs gabarito", f"{prec_gab:.1%}"),
        ("F1 (gabarito)", f"{f1:.1%}"),
        ("Precisao vs conferencia (✓ / conferidas)", f"{prec_conf:.1%}"),
    ]
    if ov:
        resumo.append(("Faixa de km IA", f"{ov[0]} a {ov[1]}"))
        resumo.append(("Faixa de km gabarito", f"{ov[2]} a {ov[3]}"))
    for met, val in resumo:
        ws3.append([met, val])
    ws3.column_dimensions["A"].width = 42
    ws3.column_dimensions["B"].width = 22

    wb.save(output)

    return {
        "output": output, "nome": nome_run, "gabarito": os.path.basename(gabarito),
        "n_det": n_det, "n_gt": n_gt, "achadas": achadas, "faltaram": faltaram,
        "n_falso": n_falso, "n_conf": n_conf, "n_real": n_real, "n_falsa": n_falsa,
        "recall": recall, "prec_gab": prec_gab, "f1": f1, "prec_conf": prec_conf,
    }


def main():
    ap = argparse.ArgumentParser(description="Gera xlsx de avaliacao de uma rodada")
    ap.add_argument("--rodada", required=True,
                    help="pasta da rodada (ex.: saidas\\PAINEL_minha_pasta)")
    ap.add_argument("--gabarito", default=GABARITO_PADRAO,
                    help="gabarito .xlsx (Comparacao.xlsx, padrao) ou .csv (inventario)")
    ap.add_argument("--km-col", default="km_abs",
                    help="coluna de km do gabarito csv (padrao: km_abs; ignorado p/ xlsx)")
    ap.add_argument("--tol-m", type=float, default=30.0,
                    help="metros max. p/ casar placa detectada ao gabarito (padrao: 30)")
    ap.add_argument("--sem-lado", action="store_true",
                    help="ignora o lado (E/D) ao casar com o gabarito")
    ap.add_argument("--output", default="",
                    help="xlsx de saida (padrao: <rodada>\\Relatorio_<nome>.xlsx)")
    args = ap.parse_args()

    r = gerar_relatorio(args.rodada, gabarito=args.gabarito, km_col=args.km_col,
                        tol_m=args.tol_m, usar_lado=not args.sem_lado, output=args.output)

    print(f"Rodada    : {os.path.abspath(args.rodada)}")
    print(f"Gabarito  : {args.gabarito} (km na faixa: {r['n_gt']})")
    print(f"IA        : {r['n_det']} placas - conferidas {r['n_conf']} "
          f"(real {r['n_real']} / falsa {r['n_falsa']})")
    print(f"Detectou {r['achadas']} - Faltaram {r['faltaram']} - Falsos {r['n_falso']}")
    print(f"Recall {r['recall']:.1%} - Precisao(gab) {r['prec_gab']:.1%} - "
          f"F1 {r['f1']:.1%} - Precisao(conf) {r['prec_conf']:.1%}")
    if r["n_gt"] == 0:
        print("[aviso] 0 placas do gabarito na faixa das fotos -- o gabarito nao cobre "
              "esse trecho/estrada (ex.: gabarito MT-361 x fotos BR-010). Use o certo.")
    print(f"Saida     : {r['output']}")


if __name__ == "__main__":
    main()
